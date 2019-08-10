"""
Read and error-check the Excel file with the music data.

The excel file is expected to have the following format:

The top row is not read.  So, can put column headings here.
Columns:
1. Don't care (its not read by this Python code).
2. Don't care (its not read by this Python code).
3. Don't care (its not read by this Python code).
4. The video frame index associated with a note.
5. Don't care (its not read by this Python code).
6. The left note value (1 = bottom A, 52 = top C).
7. The left note timing (see __init__() for list of timing strings).
8. The left note accidentals (see __init__() for list of accidental strings).
9. The right note value.
10. The right timing value.
11. The right accidental value.

If a left/right note value is 0 then this means no note for this frame index.

Gregary C. Zweigle

2019

"""

from openpyxl import load_workbook
import sys


class Data:

    def __init__(self, filename):

        wb = load_workbook(filename)
        self.sheet = wb.active
        self.index_column = 4
        self.starting_left_hand_column = 6
        self.starting_right_hand_column = 9

        self.left_hand_has_data = None
        self.left_hand_value = None
        self.left_hand_timing = None
        self.left_hand_accidental = None
        self.right_hand_has_data = None
        self.right_hand_value = None
        self.right_hand_timing = None
        self.right_hand_accidental = None

        self.row = 2

        # For validating the data read from the Excel file.
        self.note_timings = ["none", "sixteenth", "sixteenth-dot",
                             "eighth", "eighth-dot",
                             "quarter", "quarter-dot", "half", "half-dot",
                             "whole"]
        self.accidentals = ["normal", "flat", "natural", "sharp"]

    def get_music_data(self, video_index):

        self.__read_excel_file(True, video_index)
        self.__read_excel_file(False, video_index)
        self.__update_excel_row(video_index)

    def __read_excel_file(self, this_is_left_hand, video_index):

        print("Processing video index {} and comparing to next index {}".
              format(video_index,
                     self.sheet.cell(self.row, self.index_column).value))

        if video_index == self.sheet.cell(self.row, self.index_column).value:

            if this_is_left_hand:
                start_col = self.starting_left_hand_column
            else:
                start_col = self.starting_right_hand_column

            value = self.sheet.cell(self.row, start_col).value
            timing = self.sheet.cell(self.row, start_col + 1).value
            accidental = self.sheet.cell(self.row, start_col + 2).value

            self.__error_check_data(value, timing, accidental)

            if this_is_left_hand:
                if value != 0:
                    self.left_hand_has_data = True
                    self.left_hand_value = value
                    self.left_hand_timing = timing
                    self.left_hand_accidental = accidental
                else:
                    self.left_hand_has_data = False

            if not this_is_left_hand:
                if value != 0:
                    self.right_hand_has_data = True
                    self.right_hand_value = value
                    self.right_hand_timing = timing
                    self.right_hand_accidental = accidental
                else:
                    self.right_hand_has_data = False
        else:
            self.left_hand_has_data = False
            self.right_hand_has_data = False

    def __update_excel_row(self, video_index):
        if video_index == self.sheet.cell(self.row, self.index_column).value:
            self.row += 1

    def __error_check_data(self, value, timing, accidental):

        found_error = False

        if value < 0 or value > 52:
            print("In Excel file row {}, the note value [{}] is not valid.".
                  format(self.row, value))
            print("Valid notes are in range 1 (lowest A) to 52 (highest C).")
            print("Use a note value of 0 to indicate no new note.")
            found_error = True

        if timing not in self.note_timings:
            print("In Excel file row {}, the timing string [{}] is not valid.".
                  format(self.row, timing))
            print("Valid timings are:")
            print(self.note_timings)
            found_error = True

        if accidental not in self.accidentals:
            print("In Excel file row {}, accidental string [{}] is not valid.".
                  format(self.row, accidental))
            print("Valid accidentals are:")
            print(self.accidentals)
            found_error = True

        if found_error:
            sys.exit()
